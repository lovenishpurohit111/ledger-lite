from __future__ import annotations

import argparse
import asyncio
import re
import sys
from pathlib import Path

from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

from backend.app.services.pipeline import ConversionPipeline
from backend.app.services.ocr import OCRService
from backend.app.services.preprocessing import PreprocessingService


MONTH_RE = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)$", re.I)
YEAR_RE = re.compile(r"^\d{4}$")
SKIP_RE = re.compile(r"^(consolidated|standalone|rs\.?\s*crore|particulars|description|view standalone|figures in)", re.I)
TOTAL_KW = ("total", "net profit", "gross profit", "operating profit", "profit before tax", "ebitda", "net income")


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    storage = REPO_ROOT / "backend/app/storage/telegram-actions"

    table = extract_table(input_path, storage)
    if table:
        write_table_workbook(table, output_path)
        return

    pipeline = ConversionPipeline(storage)

    with input_path.open("rb") as handle:
        upload = UploadFile(file=handle, filename=input_path.name)
        result = await pipeline.run([upload])

    workbook = pipeline.workbook_path(result.job_id)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(workbook.read_bytes())


def extract_table(input_path: Path, storage: Path) -> dict | None:
    processed_dir = storage / "table-ocr"
    pages = PreprocessingService().preprocess_files([input_path], processed_dir)
    document = OCRService().extract(pages)
    for page in document.pages:
        table = build_table(page.tokens, page.text)
        if table:
            return table
    return None


def build_table(tokens, text: str) -> dict | None:
    words = [
        {
            "text": token.text,
            "bbox": token.bbox,
        }
        for token in tokens
        if token.bbox and token.text.strip()
    ]
    rows = group_rows(words)
    column_info = detect_columns(rows)
    if not column_info:
        return None

    header_index, columns = column_info
    cut_x = columns[0]["cx"] - ((columns[1]["cx"] - columns[0]["cx"]) * 0.5 if len(columns) > 1 else 60)
    out_rows = []

    for index, row_words in enumerate(rows):
        if index <= header_index:
            continue

        label = " ".join(word["text"] for word in row_words if center_x(word) < cut_x)
        label = re.sub(r"[+*|<>:]+$", "", re.sub(r"\s+", " ", label)).strip()
        if not label or len(label) < 2 or label.isdigit() or SKIP_RE.match(label):
            continue

        values = assign_values(merge_number_fragments(row_words), columns, cut_x)
        if not any(value is not None for value in values):
            continue

        out_rows.append({
            "label": label,
            "values": values,
            "row_type": classify(label),
        })

    if not out_rows:
        return None

    return {
        "statement_type": statement_type(text),
        "unit": extract_unit(text),
        "columns": [column["header"] for column in columns],
        "rows": out_rows,
    }


def group_rows(words: list[dict], y_tol: float = 10) -> list[list[dict]]:
    rows: list[dict] = []
    for word in sorted(words, key=lambda item: item["bbox"][1]):
        cy = center_y(word)
        row = next((candidate for candidate in rows if abs(candidate["cy"] - cy) < y_tol), None)
        if row:
            row["words"].append(word)
            row["cy"] = sum(center_y(item) for item in row["words"]) / len(row["words"])
        else:
            rows.append({"cy": cy, "words": [word]})
    return [sorted(row["words"], key=lambda item: item["bbox"][0]) for row in rows]


def detect_columns(rows: list[list[dict]]) -> tuple[int, list[dict]] | None:
    for index, row in enumerate(rows):
        columns = []
        i = 0
        while i < len(row):
            text = row[i]["text"]
            next_text = row[i + 1]["text"] if i + 1 < len(row) else ""
            if MONTH_RE.match(text) and YEAR_RE.match(next_text):
                columns.append({"header": f"{text} {next_text}", "cx": (row[i]["bbox"][0] + row[i + 1]["bbox"][2]) / 2})
                i += 2
                continue
            if re.match(r"^T+M[.,:]?$", text, re.I):
                columns.append({"header": "TTM", "cx": center_x(row[i])})
            i += 1

        if len(columns) >= 2:
            if not any(column["header"] == "TTM" for column in columns):
                spacing = (columns[-1]["cx"] - columns[0]["cx"]) / max(len(columns) - 1, 1)
                columns.append({"header": "TTM", "cx": columns[-1]["cx"] + spacing})
            return index, columns
    return None


def assign_values(words: list[dict], columns: list[dict], cut_x: float) -> list[float | str | None]:
    span = columns[-1]["cx"] - columns[0]["cx"]
    tolerance = max(50, (span / max(len(columns) - 1, 1)) * 0.75)
    values: list[float | str | None] = [None] * len(columns)
    used: set[int] = set()

    for word in words:
        cx = center_x(word)
        if cx < cut_x - 5:
            continue
        parsed = parse_number(word["text"])
        if parsed is None:
            continue
        best = None
        best_distance = tolerance
        for index, column in enumerate(columns):
            if index in used:
                continue
            distance = abs(cx - column["cx"])
            if distance < best_distance:
                best = index
                best_distance = distance
        if best is not None:
            values[best] = parsed
            used.add(best)
    return values


def merge_number_fragments(words: list[dict]) -> list[dict]:
    merged = []
    index = 0
    while index < len(words):
        current = words[index]
        next_word = words[index + 1] if index + 1 < len(words) else None
        if re.match(r"^-?[\d,]*\d,$", current["text"]) and next_word and re.match(r"^\d+$", next_word["text"]):
            merged.append({
                "text": current["text"] + next_word["text"],
                "bbox": [current["bbox"][0], current["bbox"][1], next_word["bbox"][2], next_word["bbox"][3]],
            })
            index += 2
        else:
            merged.append(current)
            index += 1
    return merged


def write_table_workbook(table: dict, output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = table["statement_type"][:31]
    rows = [[table["unit"], *["" for _ in table["columns"]]], ["Line Item", *table["columns"]]]
    rows.extend([[row["label"], *row["values"]] for row in table["rows"]])

    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if row_index <= 2:
                cell.font = Font(bold=True)
            if column_index > 1:
                cell.number_format = '#,##0;[Red](#,##0);-'

    sheet.freeze_panes = "B3"
    sheet.column_dimensions["A"].width = 30
    for index in range(2, len(table["columns"]) + 2):
        sheet.column_dimensions[get_column_letter(index)].width = 12

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def center_x(word: dict) -> float:
    return (word["bbox"][0] + word["bbox"][2]) / 2


def center_y(word: dict) -> float:
    return (word["bbox"][1] + word["bbox"][3]) / 2


def parse_number(value: str) -> float | str | None:
    is_percent = "%" in value
    cleaned = value.replace("%", "").replace(",", "").strip()
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return f"{number:g}%" if is_percent else number


def classify(label: str) -> str:
    lower = label.lower()
    if any(keyword in lower for keyword in TOTAL_KW):
        return "total"
    return "subtotal" if lower.startswith(("sub-total", "subtotal")) else "line_item"


def statement_type(text: str) -> str:
    lower = text.lower()
    if "balance sheet" in lower or ("assets" in lower and "liabilities" in lower):
        return "Balance Sheet"
    if "cash flow" in lower or "operating activities" in lower:
        return "Cash Flow"
    return "Profit & Loss"


def extract_unit(text: str) -> str:
    match = re.search(r"(?:Consolidated\s+)?Figures in Rs\.?\s+\w+", text, re.I)
    return match.group(0).strip() if match else "Rs. Crores"


if __name__ == "__main__":
    asyncio.run(main())
