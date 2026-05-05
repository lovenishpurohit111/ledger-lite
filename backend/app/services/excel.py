from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.models.schemas import Statement, StatementType


SHEET_NAMES = {
    StatementType.profit_and_loss: "Profit & Loss",
    StatementType.balance_sheet: "Balance Sheet",
    StatementType.cash_flow: "Cash Flow",
}


class ExcelWorkbookBuilder:
    def build(self, statements: list[Statement], output_path: Path) -> Path:
        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)

        by_type = {statement.statement_type: statement for statement in statements}
        for statement_type in [StatementType.profit_and_loss, StatementType.balance_sheet, StatementType.cash_flow]:
            statement = by_type.get(statement_type)
            sheet = workbook.create_sheet(SHEET_NAMES[statement_type])
            self._write_statement(sheet, statement_type, statement)

        workbook.save(output_path)
        return output_path

    def _write_statement(self, sheet, statement_type: StatementType, statement: Statement | None) -> None:
        sheet.freeze_panes = "A5"
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = statement.title if statement else f"{statement_type.value} Statement"
        sheet["A1"].font = Font(bold=True, size=16, color="1F2937")
        sheet["A2"] = statement.period if statement and statement.period else "Reviewed export"
        sheet["A2"].font = Font(color="6B7280")
        headers = ["Section", "Line Item", "Amount", "Confidence", "Review Notes"]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=column, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="166534")
            cell.alignment = Alignment(horizontal="center")

        rows = statement.rows if statement else []
        for row_index, row in enumerate(rows, start=5):
            sheet.cell(row=row_index, column=1, value=row.section)
            label = sheet.cell(row=row_index, column=2, value=row.label)
            label.alignment = Alignment(indent=min(row.level, 6))
            amount = sheet.cell(row=row_index, column=3, value=row.amount)
            amount.number_format = '#,##0;[Red](#,##0);-'
            sheet.cell(row=row_index, column=4, value=row.confidence).number_format = "0%"
            sheet.cell(row=row_index, column=5, value="; ".join(row.issues))

            if row.row_type in {"header", "subtotal", "total"}:
                for column in range(1, 6):
                    sheet.cell(row=row_index, column=column).font = Font(bold=True)
                    sheet.cell(row=row_index, column=column).fill = PatternFill("solid", fgColor="F3F4F6")
            if row.row_type == "total":
                for column in range(1, 6):
                    sheet.cell(row=row_index, column=column).border = Border(top=Side(style="thin", color="111827"), bottom=Side(style="double", color="111827"))

        widths = [24, 42, 18, 14, 44]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=cell.column == 5, horizontal="right" if cell.column == 3 else "left")
